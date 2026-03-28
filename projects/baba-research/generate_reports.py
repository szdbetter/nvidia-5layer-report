#!/usr/bin/env python3
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image, HRFlowable
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import os
import io

OUT_DIR = "/root/.openclaw/workspace/projects/baba-research"

# ─── Data ────────────────────────────────────────────────────────────────────

data = {
    "TTM (Dec'25)": {
        "Revenue": 1016744,
        "Gross Profit": 404374,
        "Gross Margin": 39.77,
        "SGA": 261251,
        "RD": 64510,
        "Operating Income": 79492,
        "Operating Margin": 7.82,
        "Net Income": 92810,
        "Net Margin": 8.91,
        "EPS Diluted": 37.84,
        "FCF": 94323,
        "FCF Margin": 9.28,
        "Shares": 2399,
    },
    "FY2024 (Mar'25)": {
        "Revenue": 996347,
        "Net Income": 129470,
        "EPS Diluted": 53.60,
        "FCF": 77537,
    },
    "FY2023 (Mar'24)": {
        "Revenue": 941168,
        "Net Income": 79741,
    },
}

# ─── Excel ────────────────────────────────────────────────────────────────────

def make_excel():
    wb = openpyxl.Workbook()

    orange = "FF6600"
    light_orange = "FFE0CC"
    dark_bg = "1A1A2E"
    white = "FFFFFF"
    gray = "F5F5F5"

    def hdr_style(cell, bg=orange, fg=white, bold=True, size=11):
        cell.font = Font(bold=bold, color=fg, size=size, name="Arial")
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def val_style(cell, bold=False, num_format="#,##0", align="right"):
        cell.font = Font(bold=bold, color="1A1A2E", size=10, name="Arial")
        cell.alignment = Alignment(horizontal=align, vertical="center")
        cell.number_format = num_format

    def thin_border():
        s = Side(style="thin", color="CCCCCC")
        return Border(left=s, right=s, top=s, bottom=s)

    # ── Sheet 1: Income Statement ──
    ws1 = wb.active
    ws1.title = "Income Statement"
    ws1.sheet_view.showGridLines = False

    ws1.merge_cells("A1:E1")
    ws1["A1"] = "阿里巴巴集团 BABA — Income Statement (百万人民币)"
    ws1["A1"].font = Font(bold=True, size=14, color=white, name="Arial")
    ws1["A1"].fill = PatternFill("solid", fgColor=dark_bg)
    ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 32

    headers = ["指标 / Metric", "TTM (Dec'25)", "FY2024 (Mar'25)", "FY2023 (Mar'24)", "FY2022 (Est.)"]
    for ci, h in enumerate(headers, 1):
        c = ws1.cell(row=2, column=ci, value=h)
        hdr_style(c)
        ws1.column_dimensions[get_column_letter(ci)].width = 22

    rows = [
        ("Revenue", 1016744, 996347, 941168, 853062),
        ("Gross Profit", 404374, None, None, None),
        ("Gross Margin %", 39.77, None, None, None),
        ("SG&A", 261251, None, None, None),
        ("R&D", 64510, None, None, None),
        ("Operating Income", 79492, None, None, None),
        ("Operating Margin %", 7.82, None, None, None),
        ("Net Income", 92810, 129470, 79741, None),
        ("Net Margin %", 8.91, None, None, None),
        ("EPS (Diluted) RMB", 37.84, 53.60, None, None),
        ("Shares Outstanding (M)", 2399, None, None, None),
    ]

    pct_rows = {"Gross Margin %", "Operating Margin %", "Net Margin %"}
    for ri, row in enumerate(rows, 3):
        alt = gray if ri % 2 == 0 else white
        for ci, val in enumerate(row, 1):
            c = ws1.cell(row=ri, column=ci, value=val)
            if ci == 1:
                c.font = Font(bold=True, size=10, name="Arial")
                c.fill = PatternFill("solid", fgColor=alt)
                c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            else:
                c.fill = PatternFill("solid", fgColor=alt)
                if isinstance(val, (int, float)):
                    if row[0] in pct_rows:
                        val_style(c, num_format='0.00"%"')
                    elif row[0] in ("EPS (Diluted) RMB",):
                        val_style(c, num_format="0.00")
                    else:
                        val_style(c, num_format="#,##0")
                else:
                    c.alignment = Alignment(horizontal="center")
            c.border = thin_border()
        ws1.row_dimensions[ri].height = 22

    # ── Sheet 2: Margin Analysis ──
    ws2 = wb.create_sheet("Margin Analysis")
    ws2.sheet_view.showGridLines = False

    ws2.merge_cells("A1:D1")
    ws2["A1"] = "阿里巴巴 BABA — Margin Analysis"
    ws2["A1"].font = Font(bold=True, size=14, color=white, name="Arial")
    ws2["A1"].fill = PatternFill("solid", fgColor=dark_bg)
    ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 32

    hdrs2 = ["利润率指标", "TTM (Dec'25)", "FY2024", "FY2023"]
    for ci, h in enumerate(hdrs2, 1):
        c = ws2.cell(row=2, column=ci, value=h)
        hdr_style(c)
        ws2.column_dimensions[get_column_letter(ci)].width = 22

    margin_rows = [
        ("Gross Margin %", 39.77, None, None),
        ("Operating Margin %", 7.82, None, None),
        ("Net Margin %", 8.91, 13.00, 8.47),
        ("FCF Margin %", 9.28, 7.78, None),
    ]
    for ri, row in enumerate(margin_rows, 3):
        alt = gray if ri % 2 == 0 else white
        for ci, val in enumerate(row, 1):
            c = ws2.cell(row=ri, column=ci, value=val)
            if ci == 1:
                c.font = Font(bold=True, size=10, name="Arial")
                c.fill = PatternFill("solid", fgColor=alt)
                c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            else:
                c.fill = PatternFill("solid", fgColor=alt)
                if isinstance(val, (int, float)):
                    val_style(c, num_format='0.00"%"')
                else:
                    c.alignment = Alignment(horizontal="center")
            c.border = thin_border()
        ws2.row_dimensions[ri].height = 22

    # Add bar chart for margins
    chart2 = BarChart()
    chart2.type = "col"
    chart2.title = "Margin Analysis TTM"
    chart2.style = 10
    chart2.y_axis.title = "%"
    chart2.x_axis.title = "Metric"
    chart2.shape = 4
    data_ref = Reference(ws2, min_col=2, min_row=2, max_row=6)
    cats = Reference(ws2, min_col=1, min_row=3, max_row=6)
    chart2.add_data(data_ref, titles_from_data=True)
    chart2.set_categories(cats)
    chart2.width = 16
    chart2.height = 10
    ws2.add_chart(chart2, "F3")

    # ── Sheet 3: Cash Flow ──
    ws3 = wb.create_sheet("Cash Flow")
    ws3.sheet_view.showGridLines = False

    ws3.merge_cells("A1:C1")
    ws3["A1"] = "阿里巴巴 BABA — Cash Flow (百万人民币)"
    ws3["A1"].font = Font(bold=True, size=14, color=white, name="Arial")
    ws3["A1"].fill = PatternFill("solid", fgColor=dark_bg)
    ws3["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 32

    hdrs3 = ["现金流指标", "TTM (Dec'25)", "FY2024 (Mar'25)"]
    for ci, h in enumerate(hdrs3, 1):
        c = ws3.cell(row=2, column=ci, value=h)
        hdr_style(c)
        ws3.column_dimensions[get_column_letter(ci)].width = 24

    cf_rows = [
        ("Free Cash Flow", 94323, 77537),
        ("FCF Margin %", 9.28, 7.78),
        ("FCF YoY Growth %", 21.65, None),
    ]
    for ri, row in enumerate(cf_rows, 3):
        alt = gray if ri % 2 == 0 else white
        for ci, val in enumerate(row, 1):
            c = ws3.cell(row=ri, column=ci, value=val)
            if ci == 1:
                c.font = Font(bold=True, size=10, name="Arial")
                c.fill = PatternFill("solid", fgColor=alt)
                c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            else:
                c.fill = PatternFill("solid", fgColor=alt)
                if isinstance(val, (int, float)):
                    if "%" in row[0]:
                        val_style(c, num_format='0.00"%"')
                    else:
                        val_style(c, num_format="#,##0")
                else:
                    c.alignment = Alignment(horizontal="center")
            c.border = thin_border()
        ws3.row_dimensions[ri].height = 22

    # ── Sheet 4: YoY Comparison ──
    ws4 = wb.create_sheet("YoY Comparison")
    ws4.sheet_view.showGridLines = False

    ws4.merge_cells("A1:D1")
    ws4["A1"] = "阿里巴巴 BABA — YoY Comparison"
    ws4["A1"].font = Font(bold=True, size=14, color=white, name="Arial")
    ws4["A1"].fill = PatternFill("solid", fgColor=dark_bg)
    ws4["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws4.row_dimensions[1].height = 32

    hdrs4 = ["指标", "TTM (Dec'25)", "FY2024 (Mar'25)", "YoY %"]
    for ci, h in enumerate(hdrs4, 1):
        c = ws4.cell(row=2, column=ci, value=h)
        hdr_style(c)
        ws4.column_dimensions[get_column_letter(ci)].width = 22

    yoy_rows = [
        ("Revenue", 1016744, 996347, 3.56),
        ("Net Income", 92810, 129470, -28.32),
        ("FCF", 94323, 77537, 21.65),
        ("EPS Diluted", 37.84, 53.60, -29.40),
    ]
    for ri, row in enumerate(yoy_rows, 3):
        alt = gray if ri % 2 == 0 else white
        for ci, val in enumerate(row, 1):
            c = ws4.cell(row=ri, column=ci, value=val)
            if ci == 1:
                c.font = Font(bold=True, size=10, name="Arial")
                c.fill = PatternFill("solid", fgColor=alt)
                c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            elif ci == 4:
                c.fill = PatternFill("solid", fgColor=alt)
                if isinstance(val, (int, float)):
                    val_style(c, num_format='0.00"%"')
                    if val < 0:
                        c.font = Font(color="CC0000", size=10, name="Arial")
                    else:
                        c.font = Font(color="006600", size=10, name="Arial")
            else:
                c.fill = PatternFill("solid", fgColor=alt)
                if isinstance(val, (int, float)):
                    val_style(c, num_format="#,##0")
            c.border = thin_border()
        ws4.row_dimensions[ri].height = 22

    path = os.path.join(OUT_DIR, "BABA_Financials.xlsx")
    wb.save(path)
    print(f"Excel saved: {path}")
    return path


# ─── PDF ─────────────────────────────────────────────────────────────────────

def make_charts():
    """Generate bar chart images for embedding in PDF."""
    charts = []

    # Revenue comparison
    fig, ax = plt.subplots(figsize=(7, 3.5))
    years = ["FY2023\n(Mar'24)", "FY2024\n(Mar'25)", "TTM\n(Dec'25)"]
    revenues = [941168, 996347, 1016744]
    bars = ax.bar(years, revenues, color=["#FF9944", "#FF6600", "#CC4400"], width=0.5, zorder=3)
    ax.set_title("Revenue Trend (百万人民币)", fontsize=13, fontweight="bold", pad=12)
    ax.set_ylabel("百万人民币", fontsize=10)
    ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f"{x/1e6:.2f}T" if x >= 1e6 else f"{x/1e3:.0f}B"))
    ax.grid(axis="y", linestyle="--", alpha=0.5, zorder=0)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    for bar, val in zip(bars, revenues):
        ax.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 8000,
                f"{val/1e6:.3f}T", ha="center", va="bottom", fontsize=9, fontweight="bold")
    plt.tight_layout()
    p = os.path.join(OUT_DIR, "_chart_revenue.png")
    plt.savefig(p, dpi=150, bbox_inches="tight")
    plt.close()
    charts.append(p)

    # Margin analysis
    fig, ax = plt.subplots(figsize=(7, 3.5))
    metrics = ["Gross\nMargin", "Operating\nMargin", "Net\nMargin", "FCF\nMargin"]
    margins = [39.77, 7.82, 8.91, 9.28]
    colors_list = ["#FF6600", "#FF9944", "#FFB366", "#FFC999"]
    bars = ax.bar(metrics, margins, color=colors_list, width=0.5, zorder=3)
    ax.set_title("Margin Analysis TTM (Dec'25)", fontsize=13, fontweight="bold", pad=12)
    ax.set_ylabel("%", fontsize=10)
    ax.grid(axis="y", linestyle="--", alpha=0.5, zorder=0)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    for bar, val in zip(bars, margins):
        ax.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 0.3,
                f"{val:.2f}%", ha="center", va="bottom", fontsize=9, fontweight="bold")
    plt.tight_layout()
    p = os.path.join(OUT_DIR, "_chart_margins.png")
    plt.savefig(p, dpi=150, bbox_inches="tight")
    plt.close()
    charts.append(p)

    # Net income vs FCF
    fig, ax = plt.subplots(figsize=(7, 3.5))
    x = range(2)
    labels = ["Net Income", "Free Cash Flow"]
    ttm_vals = [92810, 94323]
    fy24_vals = [129470, 77537]
    w = 0.3
    ax.bar([i - w/2 for i in x], ttm_vals, width=w, label="TTM (Dec'25)", color="#FF6600", zorder=3)
    ax.bar([i + w/2 for i in x], fy24_vals, width=w, label="FY2024 (Mar'25)", color="#FFB366", zorder=3)
    ax.set_xticks(list(x))
    ax.set_xticklabels(labels)
    ax.set_title("Net Income vs FCF (百万人民币)", fontsize=13, fontweight="bold", pad=12)
    ax.set_ylabel("百万人民币", fontsize=10)
    ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f"{x/1000:.0f}B"))
    ax.grid(axis="y", linestyle="--", alpha=0.5, zorder=0)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.legend(fontsize=9)
    plt.tight_layout()
    p = os.path.join(OUT_DIR, "_chart_ni_fcf.png")
    plt.savefig(p, dpi=150, bbox_inches="tight")
    plt.close()
    charts.append(p)

    return charts


def make_pdf(chart_paths):
    path = os.path.join(OUT_DIR, "BABA_Report.pdf")
    doc = SimpleDocTemplate(path, pagesize=A4,
                            leftMargin=2*cm, rightMargin=2*cm,
                            topMargin=2*cm, bottomMargin=2*cm)

    styles = getSampleStyleSheet()
    orange = colors.HexColor("#FF6600")
    dark = colors.HexColor("#1A1A2E")
    light_orange = colors.HexColor("#FFF3E0")

    title_style = ParagraphStyle("Title", parent=styles["Title"],
                                  fontSize=26, textColor=dark, spaceAfter=4,
                                  fontName="Helvetica-Bold", alignment=TA_CENTER)
    sub_style = ParagraphStyle("Sub", parent=styles["Normal"],
                                fontSize=13, textColor=orange, spaceAfter=2,
                                fontName="Helvetica-Bold", alignment=TA_CENTER)
    h2_style = ParagraphStyle("H2", parent=styles["Heading2"],
                               fontSize=14, textColor=dark, spaceBefore=14, spaceAfter=6,
                               fontName="Helvetica-Bold")
    body_style = ParagraphStyle("Body", parent=styles["Normal"],
                                 fontSize=10, textColor=dark, spaceAfter=4,
                                 fontName="Helvetica")
    note_style = ParagraphStyle("Note", parent=styles["Normal"],
                                 fontSize=8, textColor=colors.grey,
                                 fontName="Helvetica-Oblique")

    def section_header(text):
        return [
            HRFlowable(width="100%", thickness=2, color=orange, spaceAfter=4),
            Paragraph(text, h2_style),
        ]

    tbl_style = TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), dark),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 10),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("ALIGN", (0, 1), (0, -1), "LEFT"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, light_orange]),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 1), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
    ])

    story = []

    # ── Cover ──
    story.append(Spacer(1, 2*cm))
    story.append(Paragraph("阿里巴巴集团", title_style))
    story.append(Paragraph("Alibaba Group Holding Limited", sub_style))
    story.append(Spacer(1, 0.5*cm))
    story.append(HRFlowable(width="80%", thickness=3, color=orange, hAlign="CENTER"))
    story.append(Spacer(1, 0.5*cm))
    story.append(Paragraph("财务分析报告 Financial Analysis Report", ParagraphStyle(
        "CoverSub", parent=styles["Normal"], fontSize=16, textColor=dark,
        fontName="Helvetica-Bold", alignment=TA_CENTER)))
    story.append(Spacer(1, 0.3*cm))
    story.append(Paragraph("TTM Dec'25 | NYSE: BABA | HKEx: 9988", ParagraphStyle(
        "Ticker", parent=styles["Normal"], fontSize=11, textColor=orange,
        fontName="Helvetica", alignment=TA_CENTER)))
    story.append(Spacer(1, 0.5*cm))
    story.append(Paragraph("Generated: 2026-03-20 | Unit: 百万人民币 (RMB Millions)", note_style))
    story.append(Spacer(1, 3*cm))

    # ── Key Metrics Summary ──
    story.extend(section_header("📊 关键指标摘要 Key Metrics Summary"))
    story.append(Spacer(1, 0.2*cm))

    kpi_data = [
        ["指标", "TTM (Dec'25)", "FY2024 (Mar'25)", "YoY变化"],
        ["Revenue (百万RMB)", "1,016,744", "996,347", "+3.56% ▲"],
        ["Gross Margin", "39.77%", "N/A", "—"],
        ["Operating Margin", "7.82%", "N/A", "—"],
        ["Net Income (百万RMB)", "92,810", "129,470", "-28.32% ▼"],
        ["Net Margin", "8.91%", "13.00%", "-4.09pp ▼"],
        ["EPS Diluted (RMB)", "37.84", "53.60", "-29.40% ▼"],
        ["Free Cash Flow (百万RMB)", "94,323", "77,537", "+21.65% ▲"],
        ["FCF Margin", "9.28%", "7.78%", "+1.50pp ▲"],
        ["Diluted Shares (M)", "2,399", "—", "—"],
    ]

    kpi_style = TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), dark),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 10),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("ALIGN", (0, 1), (0, -1), "LEFT"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, light_orange]),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 1), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        # Color YoY column
        ("TEXTCOLOR", (3, 2), (3, 2), colors.HexColor("#006600")),  # Revenue
        ("TEXTCOLOR", (3, 4), (3, 4), colors.HexColor("#CC0000")),  # Net Income
        ("TEXTCOLOR", (3, 6), (3, 6), colors.HexColor("#CC0000")),  # EPS
        ("TEXTCOLOR", (3, 7), (3, 7), colors.HexColor("#006600")),  # FCF
        ("TEXTCOLOR", (3, 8), (3, 8), colors.HexColor("#006600")),  # FCF Margin
    ])

    col_widths = [5.5*cm, 3.5*cm, 3.5*cm, 3.5*cm]
    kpi_tbl = Table(kpi_data, colWidths=col_widths)
    kpi_tbl.setStyle(kpi_style)
    story.append(kpi_tbl)
    story.append(Spacer(1, 0.5*cm))

    # ── Income Statement Table ──
    story.extend(section_header("📋 收入表 Income Statement"))

    inc_data = [
        ["指标", "TTM (Dec'25)", "FY2024 (Mar'25)", "FY2023 (Mar'24)"],
        ["Revenue", "1,016,744", "996,347", "941,168"],
        ["Gross Profit", "404,374", "—", "—"],
        ["SG&A", "261,251", "—", "—"],
        ["R&D Expenses", "64,510", "—", "—"],
        ["Operating Income", "79,492", "—", "—"],
        ["Net Income", "92,810", "129,470", "79,741"],
        ["EPS Diluted (RMB)", "37.84", "53.60", "—"],
    ]
    inc_tbl = Table(inc_data, colWidths=[5.5*cm, 3.5*cm, 3.5*cm, 3.5*cm])
    inc_tbl.setStyle(tbl_style)
    story.append(inc_tbl)
    story.append(Spacer(1, 0.5*cm))

    # ── Charts ──
    story.extend(section_header("📈 图表分析 Chart Analysis"))
    for chart_p in chart_paths:
        if os.path.exists(chart_p):
            img = Image(chart_p, width=15*cm, height=7.5*cm)
            img.hAlign = "CENTER"
            story.append(img)
            story.append(Spacer(1, 0.4*cm))

    # ── Investment Highlights ──
    story.extend(section_header("💡 投资要点 Investment Highlights"))
    highlights = [
        "✅ <b>收入稳健增长</b>: TTM收入1.017万亿RMB，同比+3.56%，收入规模持续扩大。",
        "✅ <b>FCF强劲</b>: 自由现金流943亿RMB，同比+21.65%，FCF质量显著改善。",
        "⚠️ <b>净利润承压</b>: TTM净利润928亿RMB vs FY2024的1,295亿，同比下降28.3%，主要受投资收益波动影响。",
        "⚠️ <b>EPS摊薄</b>: EPS从53.60降至37.84，叠加回购仍存稀释压力，需关注股份结构变化。",
        "📌 <b>毛利率健康</b>: 毛利率39.77%，反映核心电商+云计算业务定价能力稳固。",
        "📌 <b>运营利润率偏低</b>: 7.82%的运营利润率受高费用（SG&A 261亿）拖累，运营杠杆改善空间较大。",
    ]
    for h in highlights:
        story.append(Paragraph(h, body_style))
        story.append(Spacer(1, 0.15*cm))

    story.append(Spacer(1, 0.5*cm))
    story.append(HRFlowable(width="100%", thickness=1, color=colors.grey))
    story.append(Spacer(1, 0.2*cm))
    story.append(Paragraph(
        "⚠️ 免责声明: 本报告仅供参考，不构成投资建议。数据来源为公开财报及估算，请以官方披露为准。",
        note_style))

    doc.build(story)
    print(f"PDF saved: {path}")
    return path


if __name__ == "__main__":
    excel_path = make_excel()
    chart_paths = make_charts()
    pdf_path = make_pdf(chart_paths)
    # cleanup temp charts
    for p in chart_paths:
        if os.path.exists(p):
            os.remove(p)
    print(f"\n✅ Done!")
    print(f"  Excel: {excel_path}")
    print(f"  PDF:   {pdf_path}")
